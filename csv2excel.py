# coding: utf-8

import os
import re
import csv
import xlwt
import xlrd
import xlsxwriter
import openpyxl
import pandas as pd


def is_float(text):
    pattern = re.compile(r'^[-+]?[0-9]\d*\.\d*$')
    return pattern.match(text)


def is_percent(text):
    pattern = re.compile(r'^[-+]?[0-9]\d*\.?\d* *%$')
    return pattern.match(text)


def is_int(text):
    pattern = re.compile(r'^[-+]?[0-9]\d*$')
    return pattern.match(text)


def get_short_name(full_name):
    fname = os.path.split(full_name)[-1]
    return os.path.splitext(fname)[0]


def get_excel_name(csv_name, excel_name, sheet_name):
    pname = os.path.splitext(csv_name)[0]
    if not excel_name:
        excel_name = pname + '.xls'
    if not sheet_name:
        sheet_name = os.path.splitext(os.path.split(pname)[1])[0]
    return excel_name, sheet_name


def csv2xlsx_pd1(csv_name, excel_name=None, sheet_name=None):
    f = open(csv_name, 'rt', encoding='utf-8')
    reader = csv.reader(f)

    csv_list = []
    for l in reader:
        csv_list.append(l)
    f.close()
    df = pd.DataFrame(csv_list)
    excel_name, sheet_name = get_excel_name(csv_name, excel_name, sheet_name)
    df.to_excel(excel_name, sheet_name)


# must have the same columns
def csv2xlsx_pd2(csv_name, excel_name=None, sheet_name=None):
    csv = pd.read_csv(csv_name, encoding='utf-8', header=None, sep=None)
    # print(csv)
    excel_name, sheet_name = get_excel_name(csv_name, excel_name, sheet_name)
    csv.to_excel(excel_name, sheet_name=sheet_name)


def csv2xlsx_xlwt(csv_name, excel_name=None, sheet_name=None):
    excel_name, sheet_name = get_excel_name(csv_name, excel_name, sheet_name)
    with open(csv_name, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = xlwt.Workbook()
        ws = wb.add_sheet(sheet_name, cell_overwrite_ok=True)
        # r = 0
        # for line in read:
        #     c = 0
        #     for i in line:
        #         ws.write(r, c, i)
        #         c += 1
        #     r += 1
        for r, line in enumerate(read):
            for c, cell in enumerate(line):
                ws.write(r, c, cell)
        wb.save(excel_name)


def csvmerge2excel(csv_list, excel_name=None, sheet_names=None):
    if (not excel_name) or (not excel_name.endswith('.xls')):
        excel_name = 'merge.xls'
    if not sheet_names:
        sheet_names = [get_short_name(csv_name) for csv_name in csv_list]
    fl_pat = re.compile(r'^[-+]?[0-9]\d*\.\d*$')
    percent_pat = re.compile(r'^([-+]?[0-9]\d*\.?\d*) *%$')
    int_pat = re.compile(r'^[-+]?[0-9]\d*$')
    style = xlwt.XFStyle()
    wb = xlwt.Workbook()
    for csv_name, sheet_name in zip(csv_list, sheet_names):
        with open(csv_name, 'r', encoding='utf-8') as f:
            read = csv.reader(f)
            ws = wb.add_sheet(sheet_name, cell_overwrite_ok=True)
            for r, line in enumerate(read):
                for c, cell in enumerate(line):
                    style.num_format_str = 'general'
                    if fl_pat.match(cell):
                        cell = float(cell)
                    elif int_pat.match(cell):
                        cell = int(cell)
                    else:
                        rst = percent_pat.match(cell)
                        if rst:
                            cell = float(rst.group(1)) / 100
                            style.num_format_str = '0%'
                    ws.write(r, c, cell, style)
    wb.save(excel_name)


def csv2xlsx_xlsxwriter(csv_name, excel_name=None, sheet_name=None):
    excel_name, sheet_name = get_excel_name(csv_name, excel_name, sheet_name)
    with open(csv_name, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = xlsxwriter.Workbook(excel_name)
        ws = wb.add_worksheet(sheet_name)
        for r, line in enumerate(read):
            for c, cell in enumerate(line):
                # ws.write('A1', 1)
                ws.write(r, c, cell)
        wb.close()


def csv2xlsx_openpyxl(csv_name, excel_name=None, sheet_name=None):
    excel_name, sheet_name = get_excel_name(csv_name, excel_name, sheet_name)
    with open(csv_name, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r, line in enumerate(read):
            for c, cell in enumerate(line):
                # ws['A1'] = 1
                # loc = '%s%s' % (openpyxl.utils.get_column_letter(c + 1), r + 1)
                # ws[loc] = cell
                ws.cell(column=c+1, row=r+1).value = cell
                # _ = ws.cell(column=c+1, row=r+1, value=cell)
        ws.title = sheet_name
        wb.save(excel_name)


if __name__ == '__main__':
    # csv2xlsx_pd1('D:/share/2.csv')
    csv2xlsx_xlwt('D:/share/2.csv')
    # csv2xlsx_xlsxwriter('D:/share/2.csv')
    # csv2xlsx_openpyxl('D:/share/2.csv')
